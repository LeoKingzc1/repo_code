import re
import openpyxl
import pandas as pd
import gradio as gr
from src.html_fun import *
from src.sub_match import *
from openpyxl import load_workbook


def col_index(attr):
    # if len(attr) % 2 == 0 and len(attr) <= 3:
    if len(attr) % 2 == 0:
        # for i in range(3,len(attr)):
        for i in range(len(attr)):
            # print(attr[:i], attr[i:])
            if attr[:i] == attr[i:]:
                return i
    else:
        return 0

def read_sheet(file_path, sheet_name, row):
    data = pd.read_excel(file_path, sheet_name, engine='openpyxl')
    # data = pd.read_csv(csv)
    if row != 0:
      attr = list(data.iloc[row])
      attr = [i.replace(' ', '') for i in attr]
      data.columns = attr
      data = data[row+1:]
    attr = data.columns.values.tolist()
    # print(attr)
    odd = col_index(attr)
    # print(odd)
    if odd != 0:
        data1 = data.iloc[:, :odd]
        data2 = data.iloc[:, odd:]
        df_empty = pd.concat([data1, data2], ignore_index=True)
        return df_empty
    else:
        return data

def sep_df(file_path, sheet_profit, row_p, sheet_balance, row_b, sheet_flow, row_f):
    sheet = [sheet_profit, row_p, sheet_balance, row_b, sheet_flow, row_f]
    data_p = pd.DataFrame()
    data_b = pd.DataFrame()
    data_f = pd.DataFrame()
    for i in range(0, len(sheet), 2):
        # print(i)
        sheet_name = sheet[i]

        row = int(sheet[i+1])
        if i == 0:
            data_p = read_sheet(file_path.name, sheet_name, row)
            profit_df = fin_table_re('profit', data_p)
            # data_p = profit_df
            profit_df.fillna(value='-', inplace=True)
            profit_output = df_html('利润表',profit_df)
        elif i == 2:
            data_b = read_sheet(file_path.name, sheet_name, row)
            balance_df = fin_table_re('balance', data_b)
            # data_b = balance_df
            balance_df.fillna(value='-', inplace=True)
            balance_output = df_html('资产负债表',balance_df)
        else:
            data_f = read_sheet(file_path.name, sheet_name, row)
            flow_df = fin_table_re('flow', data_f)
            # data_f = flow_df
            flow_df.fillna(value='-', inplace=True)
            flow_output = df_html('现金流量表',flow_df)
    
    with open('./html_templates/profit_template.html') as f1:
      profit_content = f1.read()
      profit_content = profit_content.replace('<!-- 在这里插入表格 -->', profit_output)

    with open('./html_templates/balance_template.html') as f2:
      balance_content = f2.read()
      balance_content = balance_content.replace('<!-- 在这里插入表格 -->', balance_output)

    with open('./html_templates/flow_template.html') as f3:
      flow_content = f3.read()
      flow_content = flow_content.replace('<!-- 在这里插入表格 -->', flow_output)
    # return content, im2
    return profit_content, balance_content, flow_content

def read_excel(file):
    workbook = load_workbook(file.name)
    sheets = []
    for i in workbook.sheetnames:
      sheets.append(i)
    # print(sheets)
    # print(type(sheets))
    return gr.update(value='', choices=sheets)

def read_df(df):
    df = pd.read_html(df,header=0)[0]
    columns = df.columns.values.tolist()
    return gr.update(value='', choices=columns)
